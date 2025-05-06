# type: ignore
# ler e alterar arquivos excel
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Define o diretório raiz e o caminho do arquivo Excel
ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "students.xlsx"

workbook: Workbook = load_workbook(WORKBOOK_PATH)
sheet_name = "My Sheet"
worksheet: Worksheet = workbook[sheet_name]

for row in worksheet.iter_rows(min_row=1):
    for cell in row:
        print(cell.value, end="\t")
    if cell.value == "John":
        worksheet.cell(cell.row, 2, 23)
    print()

workbook.save(WORKBOOK_PATH)
