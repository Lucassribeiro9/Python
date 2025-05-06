# type: ignore
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Define o diretório raiz e o caminho do arquivo Excel
ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "students.xlsx"

# Cria uma nova instância de um workbook do Excel
workbook = Workbook()

# Define o nome da planilha ativa
sheet_name = "My Sheet"
workbook.active.title = sheet_name

# Obtém a planilha ativa como um objeto Worksheet
worksheet: Worksheet = workbook[sheet_name]

# Define os cabeçalhos da planilha
worksheet.cell(1, 1, "Name")
worksheet.cell(1, 2, "Age")
worksheet.cell(1, 3, "GPA")

# Lista de estudantes com nome, idade e GPA
students = [["John", 14, 5.5], ["Jane", 13, 4.5], ["Bob", 15, 3.5], ["Alice", 12, 4.0]]

# Preenche a planilha com os dados dos estudantes
for i, students_row in enumerate(
    students, start=2
):  # start=2 para começar na linha 2 do Excel
    for j, students_col in enumerate(
        students_row, start=1
    ):  # start=1 para começar na coluna 1 do Excel
        worksheet.cell(i, j, students_col)

# Salva o workbook no caminho especificado
workbook.save(WORKBOOK_PATH)
